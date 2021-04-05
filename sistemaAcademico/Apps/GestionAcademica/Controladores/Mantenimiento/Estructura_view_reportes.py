import time
from datetime import date

import openpyxl
from django.http.response import HttpResponse, HttpResponseRedirect
from django.shortcuts import render
from django.template.loader import get_template
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from xhtml2pdf import pisa

from sistemaAcademico.Apps.GestionAcademica.Diccionario.Estructuras_tablas_conf import *
from sistemaAcademico.Apps.GestionAcademica.Diccionario.Estructuras_tablas_mant import *
from sistemaAcademico.utils import link_callback


def reporte_estudiante(request):
    if 'usuario' in request.session:
        persona = None
        if request.method == 'POST':
            usuario = ConfUsuario.objects.get(id_usuario=request.session.get('usuario'))
            campoChk = request.POST.get('check1')
            campoP = request.POST.get('campoPersona')
            combo = int(request.POST.get('combo'))
            comboR = int(request.POST.get('comboR'))
            # Aqui se realizan los filtros
            if combo == 1:
                persona = MantEstudiante.objects.filter(id_persona__apellidos__icontains=campoP)
            elif combo == 2:
                persona = MantEstudiante.objects.all()
            elif combo == 3:
                persona = MantEstudiante.objects.filter(usuario_ing=campoP)
            elif combo == 4:
                return render(request, 'sistemaAcademico/reportes/reportePersona.html')

            if comboR == 1:
                return mant_estudiante(persona, campoChk, usuario)
            elif comboR == 2:
                return ReporteEstudiante(persona, campoChk, usuario)
        return render(request, 'sistemaAcademico/reportes/reportePersona.html')
    else:
        return HttpResponseRedirect('timeout/')


def mant_estudiante(persona, campoChk3=None, usuariophh=None):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Listado de Estudiantes con Datos'
    img = openpyxl.drawing.image.Image('static/img/logo-login.png')
    img.width = 100
    img.height = 98

    ws['C2'].alignment = Alignment(horizontal="center", vertical="center")
    ws['C2'].font = Font(name='times new roman', size=12, bold=True)
    ws['C2'] = 'Fecha:'

    ws['D2'].alignment = Alignment(horizontal="center", vertical="center")
    ws['D2'].font = Font(name='times new roman', size=12)
    ws['D2'] = date.today()

    ws['C3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['C3'].font = Font(name='times new roman', size=12, bold=True)
    ws['C3'] = 'Hora:'

    ws['D3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['D3'].font = Font(name='times new roman', size=12)
    ws['D3'] = time.strftime("%H:%M")

    if campoChk3 is not None:
        usur2(ws, usuariophh)

    ws['B1'].alignment = Alignment(horizontal="center", vertical="center")
    ws['B1'].font = Font(name='times new roman', size=14, bold=True)
    ws['B1'] = 'Base De Datos De Estudiantes'

    ws['G4'].alignment = Alignment(horizontal="center", vertical="center")
    ws['G4'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['G4'].font = Font(name='times new roman', size=14, bold=True)
    ws['G4'] = 'Representante Legal'
    # Se fusionan las celdas
    ws.merge_cells('B1:L1')
    ws.merge_cells('C5:D5')
    ws.merge_cells('G4:L4')
    ws.merge_cells('H5:I5')
    # Se da tamaño a las filas
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 20
    # Se da tamaño a las columnas
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 13
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 13
    ws.column_dimensions['F'].width = 50
    ws.column_dimensions['G'].width = 13
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 13
    ws.column_dimensions['k'].width = 13
    ws.column_dimensions['L'].width = 13
    # Cabecera de la tabla
    ws['A5'].alignment = Alignment(horizontal="center", vertical="center")
    ws['A5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['A5'].fill = PatternFill(start_color='6BA3FF', end_color='6BA3FF', fill_type="solid")
    ws['A5'].font = Font(name='times new roman', size=12, bold=True)
    ws['A5'] = 'No.'

    ws['B5'].alignment = Alignment(horizontal="center", vertical="center")
    ws['B5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['B5'].fill = PatternFill(start_color='6BA3FF', end_color='6BA3FF', fill_type="solid")
    ws['B5'].font = Font(name='times new roman', size=12, bold=True)
    ws['B5'] = 'Cedula Est'

    ws['C5'].alignment = Alignment(horizontal="center", vertical="center")
    ws['C5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['C5'].fill = PatternFill(start_color='6BA3FF', end_color='6BA3FF', fill_type="solid")
    ws['C5'].font = Font(name='times new roman', size=12, bold=True)
    ws['C5'] = 'Apellidos y Nombres'

    ws['E5'].alignment = Alignment(horizontal="center", vertical="center")
    ws['E5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['E5'].fill = PatternFill(start_color='6BA3FF', end_color='6BA3FF', fill_type="solid")
    ws['E5'].font = Font(name='times new roman', size=12, bold=True)
    ws['E5'] = 'Fecha Nac'

    ws['F5'].alignment = Alignment(horizontal="center", vertical="center")
    ws['F5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['F5'].fill = PatternFill(start_color='6BA3FF', end_color='6BA3FF', fill_type="solid")
    ws['F5'].font = Font(name='times new roman', size=12, bold=True)
    ws['F5'] = 'Dirección'

    ws['G5'].alignment = Alignment(horizontal="center", vertical="center")
    ws['G5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['G5'].fill = PatternFill(start_color='6BA3FF', end_color='6BA3FF', fill_type="solid")
    ws['G5'].font = Font(name='times new roman', size=12, bold=True)
    ws['G5'] = 'Cedula Rep'

    ws['H5'].alignment = Alignment(horizontal="center", vertical="center")
    ws['H5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['H5'].fill = PatternFill(start_color='6BA3FF', end_color='6BA3FF', fill_type="solid")
    ws['H5'].font = Font(name='times new roman', size=12, bold=True)
    ws['H5'] = 'Apellidos y Nombres'

    ws['J5'].alignment = Alignment(horizontal="center", vertical="center")
    ws['J5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['J5'].fill = PatternFill(start_color='6BA3FF', end_color='6BA3FF', fill_type="solid")
    ws['J5'].font = Font(name='times new roman', size=12, bold=True)
    ws['J5'] = 'Parentesco'

    ws['K5'].alignment = Alignment(horizontal="center", vertical="center")
    ws['K5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['K5'].fill = PatternFill(start_color='6BA3FF', end_color='6BA3FF', fill_type="solid")
    ws['K5'].font = Font(name='times new roman', size=12, bold=True)
    ws['K5'] = 'Num Conv'

    ws['L5'].alignment = Alignment(horizontal="center", vertical="center")
    ws['L5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['L5'].fill = PatternFill(start_color='6BA3FF', end_color='6BA3FF', fill_type="solid")
    ws['L5'].font = Font(name='times new roman', size=12, bold=True)
    ws['L5'] = 'Num Celu'

    # presentacion de los datos consultados en al base
    controlador = 6
    cont = 1
    for mant in persona:
        ws.merge_cells(f"C{controlador}:D{controlador}")
        ws.merge_cells(f"H{controlador}:I{controlador}")
        ws.cell(row=controlador, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=1).border = Border(left=Side(border_style="thin"),
                                                           right=Side(border_style="thin"),
                                                           top=Side(border_style="thin"),
                                                           bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=1).font = Font(name='times new roman', size=11)
        ws.cell(row=controlador, column=1).value = cont

        ws.cell(row=controlador, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=2).border = Border(left=Side(border_style="thin"),
                                                           right=Side(border_style="thin"),
                                                           top=Side(border_style="thin"),
                                                           bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=2).font = Font(name='times new roman', size=11)
        ws.cell(row=controlador, column=2).value = mant.id_persona.identificacion

        ws.cell(row=controlador, column=3).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=3).border = Border(left=Side(border_style="thin"),
                                                           right=Side(border_style="thin"),
                                                           top=Side(border_style="thin"),
                                                           bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=3).font = Font(name='times new roman', size=11)
        ws.cell(row=controlador, column=3).value = mant.id_persona.nombres + ' ' + mant.id_persona.apellidos

        ws.cell(row=controlador, column=5).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=5).border = Border(left=Side(border_style="thin"),
                                                           right=Side(border_style="thin"),
                                                           top=Side(border_style="thin"),
                                                           bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=5).font = Font(name='times new roman', size=11)
        ws.cell(row=controlador, column=5).value = mant.id_persona.fecha_de_nacimiento

        ws.cell(row=controlador, column=6).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=6).border = Border(left=Side(border_style="thin"),
                                                           right=Side(border_style="thin"),
                                                           top=Side(border_style="thin"),
                                                           bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=6).font = Font(name='times new roman', size=11)
        ws.cell(row=controlador, column=6).value = mant.id_persona.direccion

        ws.cell(row=controlador, column=7).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=7).border = Border(left=Side(border_style="thin"),
                                                           right=Side(border_style="thin"),
                                                           top=Side(border_style="thin"),
                                                           bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=7).font = Font(name='times new roman', size=11)
        ws.cell(row=controlador, column=7).value = mant.id_persona.ridentificacion

        ws.cell(row=controlador, column=8).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=8).border = Border(left=Side(border_style="thin"),
                                                           right=Side(border_style="thin"),
                                                           top=Side(border_style="thin"),
                                                           bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=8).font = Font(name='times new roman', size=11)
        ws.cell(row=controlador, column=8).value = mant.id_persona.rapellidos + ' ' + mant.id_persona.rnombres

        ws.cell(row=controlador, column=10).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=10).border = Border(left=Side(border_style="thin"),
                                                            right=Side(border_style="thin"),
                                                            top=Side(border_style="thin"),
                                                            bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=10).font = Font(name='times new roman', size=11)
        ws.cell(row=controlador, column=10).value = mant.id_persona.tipo_parentesco

        ws.cell(row=controlador, column=11).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=11).border = Border(left=Side(border_style="thin"),
                                                            right=Side(border_style="thin"),
                                                            top=Side(border_style="thin"),
                                                            bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=11).font = Font(name='times new roman', size=11)
        ws.cell(row=controlador, column=11).value = mant.id_persona.rtelefono_trabajo

        ws.cell(row=controlador, column=12).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=12).border = Border(left=Side(border_style="thin"),
                                                            right=Side(border_style="thin"),
                                                            top=Side(border_style="thin"),
                                                            bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=12).font = Font(name='times new roman', size=11)
        ws.cell(row=controlador, column=12).value = mant.id_persona.rtelefono

        controlador += 1
        cont += 1

    # establecer el nombre de mi archivo
    nombre_archivo = "ReporteEstudiantes.xlsx"
    # Definir tipo de respuesta que va a dar
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename = {0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    # Se añade la imagen que se encuentra en la esquina de la hoja
    ws.add_image(img, 'A1')
    wb.save('logo-login.xlsx')
    wb.save(response)
    return response


# En el caso de seleccionar la casilla de mostrar usuario, se mostrara el usuario en el documento
def usur2(ws, usuario):
    ws['C4'].alignment = Alignment(horizontal="center", vertical="center")
    ws['C4'].font = Font(name='times new roman', size=12, bold=True)
    ws['C4'] = 'Usuario: '

    ws['D4'].alignment = Alignment(horizontal="center", vertical="center")
    ws['D4'].font = Font(name='times new roman', size=12)
    ws['D4'] = ' {0}'.format(usuario)


def ReporteEstudiante(persona, campoChk=None, usuarioph=None):
    template_path = 'sistemaAcademico/DiseñoReporte/DiseñoEstudiante.html'
    response = HttpResponse(content_type='application/pdf')
    context = dict()
    context['fecha_actual'] = date.today()
    context['hora_actual'] = time.strftime("%H:%M")
    response['Content-Disposition'] = 'attachment; filename=ReporteEstudiante.pdf'
    context['lista_estudiante'] = persona
    if campoChk is not None:
        estu(context, usuarioph)

    template = get_template(template_path)
    html = template.render(context)
    pisaStatus = pisa.CreatePDF(html, dest=response, link_callback=link_callback)
    if pisaStatus.err:
        return HttpResponse('we had some errors <pre>' + html + '</pre>')
    return response


def estu(context, usuario):
    context['nombre_usuario'] = usuario


##################################################################################################################

def reporte_empleado(request):
    if 'usuario' in request.session:
        empleado = None
        if request.method == 'POST':
            usuario = ConfUsuario.objects.get(id_usuario=request.session.get('usuario'))
            campoChk = request.POST.get('check1')
            campoP = request.POST.get('campoEmpleado')
            combo = int(request.POST.get('combo'))
            comboR = int(request.POST.get('comboR'))
            if combo == 1:
                empleado = MantEmpleado.objects.filter(id_persona__apellidos__icontains=campoP)
            elif combo == 3:
                empleado = MantEmpleado.objects.filter(id_persona__id_genr_tipo_usuario__nombre=campoP)
            elif combo == 2:
                empleado = MantEmpleado.objects.all()
            elif combo == 4:
                return render(request, 'sistemaAcademico/reportes/reporteEmpleado.html')

            if comboR == 1:
                return mant_empleado(empleado, campoChk, usuario)
            elif comboR == 2:
                return ReporteEmpleado(empleado, campoChk, usuario)
        return render(request, 'sistemaAcademico/reportes/reporteEmpleado.html')
    else:
        return HttpResponseRedirect('timeout/')


def mant_empleado(empleado, campoChk=None, usuarioph=None):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Ficha de Datos Docentes'
    img = openpyxl.drawing.image.Image('static/img/logo-login.png')
    img.width = 100
    img.height = 98

    # ---------------------------------para darle diseño a mi titulo en la hoja-----------------------------------------
    ws['C2'].alignment = Alignment(horizontal="center", vertical="center")
    ws['C2'].font = Font(name='times new roman', size=12, bold=True)
    ws['C2'] = 'Fecha:'

    ws['D2'].alignment = Alignment(horizontal="center", vertical="center")
    ws['D2'].font = Font(name='times new roman', size=12)
    ws['D2'] = date.today()

    ws['C3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['C3'].font = Font(name='times new roman', size=12, bold=True)
    ws['C3'] = 'Hora:'

    ws['D3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['D3'].font = Font(name='times new roman', size=12)
    ws['D3'] = time.strftime("%H:%M")

    if campoChk is not None:
        usur(ws, usuarioph)

    ws['B1'].alignment = Alignment(horizontal="center", vertical="center")
    ws['B1'].font = Font(name='times new roman', size=14, bold=True)
    ws['B1'] = 'Base de Datos Empleados'
    # ---------------------------------------cambiar caracteristicas de las celdas--------------------------------------
    ws.merge_cells('B1:AD1')
    ws.merge_cells('E5:F5')

    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 20
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 13
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 13
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 30
    ws.column_dimensions['k'].width = 25
    ws.column_dimensions['L'].width = 20
    ws.column_dimensions['M'].width = 20
    ws.column_dimensions['N'].width = 20
    ws.column_dimensions['O'].width = 20
    ws.column_dimensions['P'].width = 20
    ws.column_dimensions['Q'].width = 20
    ws.column_dimensions['R'].width = 20
    ws.column_dimensions['S'].width = 20
    ws.column_dimensions['T'].width = 20
    ws.column_dimensions['U'].width = 20
    ws.column_dimensions['V'].width = 20
    ws.column_dimensions['W'].width = 25
    ws.column_dimensions['X'].width = 20
    ws.column_dimensions['Y'].width = 20
    ws.column_dimensions['Z'].width = 20
    ws.column_dimensions['AA'].width = 20
    ws.column_dimensions['AB'].width = 20
    ws.column_dimensions['AC'].width = 20
    ws.column_dimensions['AD'].width = 20

    # ws.column_dimensions['D'].width = 20
    # ----------------------------------------------------darle diseño a mi cabecera------------------------------------

    ws['A5'].alignment = Alignment(horizontal="center", vertical="center")
    ws['A5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['A5'].fill = PatternFill(start_color='3380FF', end_color='3380FF', fill_type="solid")
    ws['A5'].font = Font(name='times new roman', size=12, bold=True)
    ws['A5'] = 'No.'

    ws['B5'].alignment = Alignment(horizontal="center", vertical="center")
    ws['B5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['B5'].fill = PatternFill(start_color='3380FF', end_color='3380FF', fill_type="solid")
    ws['B5'].font = Font(name='times new roman', size=12, bold=True)
    ws['B5'] = 'Estado'

    ws['C5'].alignment = Alignment(horizontal="center", vertical="center")
    ws['C5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['C5'].fill = PatternFill(start_color='3380FF', end_color='3380FF', fill_type="solid")
    ws['C5'].font = Font(name='times new roman', size=12, bold=True)
    ws['C5'] = 'Cargo'

    ws['D5'].alignment = Alignment(horizontal="center", vertical="center")
    ws['D5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['D5'].fill = PatternFill(start_color='3380FF', end_color='3380FF', fill_type="solid")
    ws['D5'].font = Font(name='times new roman', size=12, bold=True)
    ws['D5'] = 'Cedula'

    ws['E5'].alignment = Alignment(horizontal="center", vertical="center")
    ws['E5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['E5'].fill = PatternFill(start_color='3380FF', end_color='3380FF', fill_type="solid")
    ws['E5'].font = Font(name='times new roman', size=12, bold=True)
    ws['E5'] = 'Apellidos y Nombres'

    ws['G5'].alignment = Alignment(horizontal="center", vertical="center")
    ws['G5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['G5'].fill = PatternFill(start_color='3380FF', end_color='3380FF', fill_type="solid")
    ws['G5'].font = Font(name='times new roman', size=12, bold=True)
    ws['G5'] = 'Fecha Nac'

    ws['H5'].alignment = Alignment(horizontal="center", vertical="center")
    ws['H5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['H5'].fill = PatternFill(start_color='3380FF', end_color='3380FF', fill_type="solid")
    ws['H5'].font = Font(name='times new roman', size=12, bold=True)
    ws['H5'] = 'Sexo'

    ws['I5'].alignment = Alignment(horizontal="center", vertical="center")
    ws['I5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['I5'].fill = PatternFill(start_color='3380FF', end_color='3380FF', fill_type="solid")
    ws['I5'].font = Font(name='times new roman', size=12, bold=True)
    ws['I5'] = 'Nombramiento'

    ws['J5'].alignment = Alignment(horizontal="center", vertical="center")
    ws['J5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['J5'].fill = PatternFill(start_color='3380FF', end_color='3380FF', fill_type="solid")
    ws['J5'].font = Font(name='times new roman', size=12, bold=True)
    ws['J5'] = 'Numero ultima accion de personal'

    ws['K5'].alignment = Alignment(horizontal="center", vertical="center")
    ws['K5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['K5'].fill = PatternFill(start_color='3380FF', end_color='3380FF', fill_type="solid")
    ws['K5'].font = Font(name='times new roman', size=12, bold=True)
    ws['K5'] = 'Fecha ultima accion de personal'

    ws['L5'].alignment = Alignment(horizontal="center", vertical="center")
    ws['L5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['L5'].fill = PatternFill(start_color='3380FF', end_color='3380FF', fill_type="solid")
    ws['L5'].font = Font(name='times new roman', size=12, bold=True)
    ws['L5'] = 'Etnia'

    ws['M5'].alignment = Alignment(horizontal="center", vertical="center")
    ws['M5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['M5'].fill = PatternFill(start_color='3380FF', end_color='3380FF', fill_type="solid")
    ws['M5'].font = Font(name='times new roman', size=12, bold=True)
    ws['M5'] = 'Discapacidad'

    ws['N5'].alignment = Alignment(horizontal="center", vertical="center")
    ws['N5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['N5'].fill = PatternFill(start_color='3380FF', end_color='3380FF', fill_type="solid")
    ws['N5'].font = Font(name='times new roman', size=12, bold=True)
    ws['N5'] = 'Tipo de Discapacidad'

    ws['O5'].alignment = Alignment(horizontal="center", vertical="center")
    ws['O5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['O5'].fill = PatternFill(start_color='3380FF', end_color='3380FF', fill_type="solid")
    ws['O5'].font = Font(name='times new roman', size=12, bold=True)
    ws['O5'] = 'Estado civil'

    ws['P5'].alignment = Alignment(horizontal="center", vertical="center")
    ws['P5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['P5'].fill = PatternFill(start_color='3380FF', end_color='3380FF', fill_type="solid")
    ws['P5'].font = Font(name='times new roman', size=12, bold=True)
    ws['P5'] = 'Direccion'

    ws['Q5'].alignment = Alignment(horizontal="center", vertical="center")
    ws['Q5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['Q5'].fill = PatternFill(start_color='3380FF', end_color='3380FF', fill_type="solid")
    ws['Q5'].font = Font(name='times new roman', size=12, bold=True)
    ws['Q5'] = 'N convencional'

    ws['R5'].alignment = Alignment(horizontal="center", vertical="center")
    ws['R5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['R5'].fill = PatternFill(start_color='3380FF', end_color='3380FF', fill_type="solid")
    ws['R5'].font = Font(name='times new roman', size=12, bold=True)
    ws['R5'] = 'N celular'

    ws['S5'].alignment = Alignment(horizontal="center", vertical="center")
    ws['S5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['S5'].fill = PatternFill(start_color='3380FF', end_color='3380FF', fill_type="solid")
    ws['S5'].font = Font(name='times new roman', size=12, bold=True)
    ws['S5'] = 'Correo Personal'

    ws['T5'].alignment = Alignment(horizontal="center", vertical="center")
    ws['T5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['T5'].fill = PatternFill(start_color='3380FF', end_color='3380FF', fill_type="solid")
    ws['T5'].font = Font(name='times new roman', size=12, bold=True)
    ws['T5'] = 'Titulo 4° nivel'

    ws['U5'].alignment = Alignment(horizontal="center", vertical="center")
    ws['U5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['U5'].fill = PatternFill(start_color='3380FF', end_color='3380FF', fill_type="solid")
    ws['U5'].font = Font(name='times new roman', size=12, bold=True)
    ws['U5'] = '1er. Titulo 3° Nivel'

    ws['V5'].alignment = Alignment(horizontal="center", vertical="center")
    ws['V5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['V5'].fill = PatternFill(start_color='3380FF', end_color='3380FF', fill_type="solid")
    ws['V5'].font = Font(name='times new roman', size=12, bold=True)
    ws['V5'] = '2do. Titulo 3° Nivel'

    ws['W5'].alignment = Alignment(horizontal="center", vertical="center")
    ws['W5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['W5'].fill = PatternFill(start_color='3380FF', end_color='3380FF', fill_type="solid")
    ws['W5'].font = Font(name='times new roman', size=12, bold=True)
    ws['W5'] = 'Titulo Tecnivo - Tecnologo'

    ws['X5'].alignment = Alignment(horizontal="center", vertical="center")
    ws['X5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['X5'].fill = PatternFill(start_color='3380FF', end_color='3380FF', fill_type="solid")
    ws['X5'].font = Font(name='times new roman', size=12, bold=True)
    ws['X5'] = 'Fecha Ingr Magis'

    ws['Y5'].alignment = Alignment(horizontal="center", vertical="center")
    ws['Y5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['Y5'].fill = PatternFill(start_color='3380FF', end_color='3380FF', fill_type="solid")
    ws['Y5'].font = Font(name='times new roman', size=12, bold=True)
    ws['Y5'] = 'Correo Institucional'

    ws['Z5'].alignment = Alignment(horizontal="center", vertical="center")
    ws['Z5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['Z5'].fill = PatternFill(start_color='3380FF', end_color='3380FF', fill_type="solid")
    ws['Z5'].font = Font(name='times new roman', size=12, bold=True)
    ws['Z5'] = 'Categoria Docente'

    ws['AA5'].alignment = Alignment(horizontal="center", vertical="center")
    ws['AA5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                              top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['AA5'].fill = PatternFill(start_color='3380FF', end_color='3380FF', fill_type="solid")
    ws['AA5'].font = Font(name='times new roman', size=12, bold=True)
    ws['AA5'] = 'Fecha Ingreso a JRA'

    ws['AB5'].alignment = Alignment(horizontal="center", vertical="center")
    ws['AB5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                              top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['AB5'].fill = PatternFill(start_color='3380FF', end_color='3380FF', fill_type="solid")
    ws['AB5'].font = Font(name='times new roman', size=12, bold=True)
    ws['AB5'] = 'Motivo Ingreso'

    ws['AC5'].alignment = Alignment(horizontal="center", vertical="center")
    ws['AC5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                              top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['AC5'].fill = PatternFill(start_color='3380FF', end_color='3380FF', fill_type="solid")
    ws['AC5'].font = Font(name='times new roman', size=12, bold=True)
    ws['AC5'] = 'Fecha Salida JRA'

    ws['AD5'].alignment = Alignment(horizontal="center", vertical="center")
    ws['AD5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                              top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['AD5'].fill = PatternFill(start_color='3380FF', end_color='3380FF', fill_type="solid")
    ws['AD5'].font = Font(name='times new roman', size=12, bold=True)
    ws['AD5'] = 'Motivo Salida'

    # ---------------------pintar datos en excel y EXPORTAR DATOS DE LA BD------------------------------------
    controlador = 6
    cont = 1
    for empe in empleado:
        ws.merge_cells(f"E{controlador}:F{controlador}")
        ws.cell(row=controlador, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=1).border = Border(left=Side(border_style="thin"),
                                                           right=Side(border_style="thin"),
                                                           top=Side(border_style="thin"),
                                                           bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=1).font = Font(name='times new roman', size=11)
        ws.cell(row=controlador, column=1).value = cont

        ws.cell(row=controlador, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=2).border = Border(left=Side(border_style="thin"),
                                                           right=Side(border_style="thin"),
                                                           top=Side(border_style="thin"),
                                                           bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=2).font = Font(name='times new roman', size=11)
        ws.cell(row=controlador, column=2).value = empe.id_persona.estado.nombre

        ws.cell(row=controlador, column=3).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=3).border = Border(left=Side(border_style="thin"),
                                                           right=Side(border_style="thin"),
                                                           top=Side(border_style="thin"),
                                                           bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=3).font = Font(name='times new roman', size=11)
        ws.cell(row=controlador, column=3).value = empe.id_persona.id_genr_tipo_usuario.nombre

        ws.cell(row=controlador, column=4).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=4).border = Border(left=Side(border_style="thin"),
                                                           right=Side(border_style="thin"),
                                                           top=Side(border_style="thin"),
                                                           bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=4).font = Font(name='times new roman', size=11)
        ws.cell(row=controlador, column=4).value = empe.id_persona.identificacion

        ws.cell(row=controlador, column=5).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=5).border = Border(left=Side(border_style="thin"),
                                                           right=Side(border_style="thin"),
                                                           top=Side(border_style="thin"),
                                                           bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=5).font = Font(name='times new roman', size=11)
        ws.cell(row=controlador, column=5).value = empe.id_persona.apellidos + ' ' + empe.id_persona.nombres

        ws.cell(row=controlador, column=7).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=7).border = Border(left=Side(border_style="thin"),
                                                           right=Side(border_style="thin"),
                                                           top=Side(border_style="thin"),
                                                           bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=7).font = Font(name='times new roman', size=11)
        ws.cell(row=controlador, column=7).value = empe.id_persona.fecha_de_nacimiento

        ws.cell(row=controlador, column=8).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=8).border = Border(left=Side(border_style="thin"),
                                                           right=Side(border_style="thin"),
                                                           top=Side(border_style="thin"),
                                                           bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=8).font = Font(name='times new roman', size=11)
        ws.cell(row=controlador, column=8).value = empe.id_persona.id_genr_genero.nombre

        ws.cell(row=controlador, column=9).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=9).border = Border(left=Side(border_style="thin"),
                                                           right=Side(border_style="thin"),
                                                           top=Side(border_style="thin"),
                                                           bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=9).font = Font(name='times new roman', size=11)
        ws.cell(row=controlador, column=9).value = 0

        ws.cell(row=controlador, column=10).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=10).border = Border(left=Side(border_style="thin"),
                                                            right=Side(border_style="thin"),
                                                            top=Side(border_style="thin"),
                                                            bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=10).font = Font(name='times new roman', size=11)
        ws.cell(row=controlador, column=10).value = 0

        ws.cell(row=controlador, column=11).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=11).border = Border(left=Side(border_style="thin"),
                                                            right=Side(border_style="thin"),
                                                            top=Side(border_style="thin"),
                                                            bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=11).font = Font(name='times new roman', size=11)
        ws.cell(row=controlador, column=11).value = 0

        ws.cell(row=controlador, column=12).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=12).border = Border(left=Side(border_style="thin"),
                                                            right=Side(border_style="thin"),
                                                            top=Side(border_style="thin"),
                                                            bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=12).font = Font(name='times new roman', size=11)
        ws.cell(row=controlador, column=12).value = empe.id_persona.id_genr_etnia.nombre

        ws.cell(row=controlador, column=13).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=13).border = Border(left=Side(border_style="thin"),
                                                            right=Side(border_style="thin"),
                                                            top=Side(border_style="thin"),
                                                            bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=13).font = Font(name='times new roman', size=11)
        ws.cell(row=controlador, column=13).value = empe.id_persona.discapacidad

        ws.cell(row=controlador, column=14).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=14).border = Border(left=Side(border_style="thin"),
                                                            right=Side(border_style="thin"),
                                                            top=Side(border_style="thin"),
                                                            bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=14).font = Font(name='times new roman', size=11)
        ws.cell(row=controlador, column=14).value = 0

        ws.cell(row=controlador, column=15).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=15).border = Border(left=Side(border_style="thin"),
                                                            right=Side(border_style="thin"),
                                                            top=Side(border_style="thin"),
                                                            bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=15).font = Font(name='times new roman', size=11)
        ws.cell(row=controlador, column=15).value = empe.id_persona.id_genr_estado_civil.nombre

        ws.cell(row=controlador, column=16).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=16).border = Border(left=Side(border_style="thin"),
                                                            right=Side(border_style="thin"),
                                                            top=Side(border_style="thin"),
                                                            bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=16).font = Font(name='times new roman', size=11)
        ws.cell(row=controlador, column=16).value = empe.id_persona.direccion

        ws.cell(row=controlador, column=17).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=17).border = Border(left=Side(border_style="thin"),
                                                            right=Side(border_style="thin"),
                                                            top=Side(border_style="thin"),
                                                            bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=17).font = Font(name='times new roman', size=11)
        ws.cell(row=controlador, column=17).value = empe.id_persona.telefono

        ws.cell(row=controlador, column=18).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=18).border = Border(left=Side(border_style="thin"),
                                                            right=Side(border_style="thin"),
                                                            top=Side(border_style="thin"),
                                                            bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=18).font = Font(name='times new roman', size=11)
        ws.cell(row=controlador, column=18).value = empe.id_persona.celular

        ws.cell(row=controlador, column=19).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=19).border = Border(left=Side(border_style="thin"),
                                                            right=Side(border_style="thin"),
                                                            top=Side(border_style="thin"),
                                                            bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=19).font = Font(name='times new roman', size=11)
        ws.cell(row=controlador, column=19).value = 0

        ws.cell(row=controlador, column=20).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=20).border = Border(left=Side(border_style="thin"),
                                                            right=Side(border_style="thin"),
                                                            top=Side(border_style="thin"),
                                                            bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=20).font = Font(name='times new roman', size=11)
        ws.cell(row=controlador, column=20).value = 0

        ws.cell(row=controlador, column=21).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=21).border = Border(left=Side(border_style="thin"),
                                                            right=Side(border_style="thin"),
                                                            top=Side(border_style="thin"),
                                                            bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=21).font = Font(name='times new roman', size=11)
        ws.cell(row=controlador, column=21).value = 0

        ws.cell(row=controlador, column=22).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=22).border = Border(left=Side(border_style="thin"),
                                                            right=Side(border_style="thin"),
                                                            top=Side(border_style="thin"),
                                                            bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=22).font = Font(name='times new roman', size=11)
        ws.cell(row=controlador, column=22).value = 0

        ws.cell(row=controlador, column=23).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=23).border = Border(left=Side(border_style="thin"),
                                                            right=Side(border_style="thin"),
                                                            top=Side(border_style="thin"),
                                                            bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=23).font = Font(name='times new roman', size=11)
        ws.cell(row=controlador, column=23).value = 0

        ws.cell(row=controlador, column=24).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=24).border = Border(left=Side(border_style="thin"),
                                                            right=Side(border_style="thin"),
                                                            top=Side(border_style="thin"),
                                                            bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=24).font = Font(name='times new roman', size=11)
        ws.cell(row=controlador, column=24).value = 0

        ws.cell(row=controlador, column=25).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=25).border = Border(left=Side(border_style="thin"),
                                                            right=Side(border_style="thin"),
                                                            top=Side(border_style="thin"),
                                                            bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=25).font = Font(name='times new roman', size=11)
        ws.cell(row=controlador, column=25).value = 0

        ws.cell(row=controlador, column=26).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=26).border = Border(left=Side(border_style="thin"),
                                                            right=Side(border_style="thin"),
                                                            top=Side(border_style="thin"),
                                                            bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=26).font = Font(name='times new roman', size=11)
        ws.cell(row=controlador, column=26).value = 0

        ws.cell(row=controlador, column=27).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=27).border = Border(left=Side(border_style="thin"),
                                                            right=Side(border_style="thin"),
                                                            top=Side(border_style="thin"),
                                                            bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=27).font = Font(name='times new roman', size=11)
        ws.cell(row=controlador, column=27).value = 0

        ws.cell(row=controlador, column=28).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=28).border = Border(left=Side(border_style="thin"),
                                                            right=Side(border_style="thin"),
                                                            top=Side(border_style="thin"),
                                                            bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=28).font = Font(name='times new roman', size=11)
        ws.cell(row=controlador, column=28).value = 0

        ws.cell(row=controlador, column=29).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=29).border = Border(left=Side(border_style="thin"),
                                                            right=Side(border_style="thin"),
                                                            top=Side(border_style="thin"),
                                                            bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=29).font = Font(name='times new roman', size=11)
        ws.cell(row=controlador, column=29).value = 0

        ws.cell(row=controlador, column=30).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=30).border = Border(left=Side(border_style="thin"),
                                                            right=Side(border_style="thin"),
                                                            top=Side(border_style="thin"),
                                                            bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=30).font = Font(name='times new roman', size=11)
        ws.cell(row=controlador, column=30).value = 0

        controlador += 1
        cont += 1

    # establecer el nombre de mi archivo
    nombre_archivo = "ReporteEmpleados.xlsx"
    # Definir tipo de respuesta que va a dar
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename = {0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido

    ws.add_image(img, 'A1')
    wb.save('logo-login.xlsx')
    wb.save(response)
    return response


def usur(ws, usuario):
    ws['C4'].alignment = Alignment(horizontal="center", vertical="center")
    ws['C4'].font = Font(name='times new roman', size=12, bold=True)
    ws['C4'] = 'Usuario: '

    ws['D4'].alignment = Alignment(horizontal="center", vertical="center")
    ws['D4'].font = Font(name='times new roman', size=12)
    ws['D4'] = ' {0}'.format(usuario)


def ReporteEmpleado(empleado, campoChk=None, usuarioph=None):
    template_path = 'sistemaAcademico/DiseñoReporte/DiseñoEmpleado.html'
    response = HttpResponse(content_type='application/pdf')
    context = dict()
    context['fecha_actual'] = date.today()
    context['hora_actual'] = time.strftime("%H:%M")
    response['Content-Disposition'] = 'attachment; filename=ReporteEmpleado.pdf'
    context['lista_empleado'] = empleado
    if campoChk is not None:
        usu(context, usuarioph)

    template = get_template(template_path)
    html = template.render(context)
    pisaStatus = pisa.CreatePDF(html, dest=response, link_callback=link_callback)
    if pisaStatus.err:
        return HttpResponse('we had some errors <pre>' + html + '</pre>')
    return response


def usu(context, usuario):
    context['nombre_usuario'] = usuario
