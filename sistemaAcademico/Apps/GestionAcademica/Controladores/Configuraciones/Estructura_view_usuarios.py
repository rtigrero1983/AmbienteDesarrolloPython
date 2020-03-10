import socket
from django.http import HttpResponseRedirect, HttpResponse
from django.shortcuts import render, redirect
from django.utils import timezone
import hashlib
import os
from django.views.decorators.cache import cache_page
from sistemaAcademico.Apps.GestionAcademica.Diccionario.Estructuras_tablas_conf import *
from sistemaAcademico.Apps.GestionAcademica.Diccionario.Estructuras_tablas_genr import *
from sistemaAcademico.Apps.GestionAcademica import forms
from django.views.generic import ListView, CreateView, UpdateView
from django.urls import reverse_lazy
from sistemaAcademico.Apps.GestionAcademica.Forms.Configuracion.forms_configuraciones import UsuarioModelForm,UsuarioeditModelForm
from django.urls import reverse


class Usuarios(ListView):
    model = ConfUsuario
    template_name = 'sistemaAcademico/Configuraciones/Usuarios/usuario.html'
    queryset = ConfUsuario.objects.filter(id_genr_estado=97).select_related(
        'id_persona', 'id_genr_tipo_usuario')
    context_object_name = 'lista_usuarios'


class CreateUsuario(CreateView):
    model=ConfUsuario
    form_class = UsuarioModelForm
    context_object_name = 'm'
    template_name = 'sistemaAcademico/Configuraciones/Usuarios/crear-usuario.html'
    success_url = reverse_lazy('Academico:usuarios')

    def get_context_data(self, **kwargs):
        context = super(CreateUsuario, self).get_context_data(**kwargs)
        context['rol'] = ConfRol.objects.all()
        return context

    def post(self, request, *args, **kargs):
        self.object = self.get_object
        form = self.form_class(request.POST)
        if form.is_valid():
            usuario = form.save()
            var_contra = usuario.clave
            h = hashlib.new("sha1")
            var_contra = str.encode(var_contra)
            h.update(var_contra)
            usuario.clave = h.hexdigest()
            usuario.save()
            return redirect(self.get_success_url())
        else:
            return self.render_to_response(self.get_context_data(form=form))

class UpdateUsuario(UpdateView):
    model = ConfUsuario
    form_class = UsuarioeditModelForm
    context_object_name = 'n'
    template_name = 'sistemaAcademico/Configuraciones/Usuarios/editar-usuario.html'
    success_url = reverse_lazy('Academico:usuarios')


def eliminar_usuario(request, id):
    usuarios = ConfUsuario.objects.get(id_usuario=id)
    inactivo = GenrGeneral.objects.get(idgenr_general=98)
    if request.method == 'POST':
        usuarios.id_genr_estado = inactivo
        usuarios.save()
        return redirect('Academico:usuarios')
    return render(request, 'sistemaAcademico/Configuraciones/Usuarios/eliminar.html', {'usuario': usuarios})

#------------- Seccion de reportes--------------------------------


from django.shortcuts import render
from django.views.generic import TemplateView
from openpyxl import Workbook #nos permite crear libro de trabajo en excel
import datetime
import time
from openpyxl.styles import Alignment,Border,Font,PatternFill,Side
from openpyxl.drawing.image import Image
from django.http.response import HttpResponse, HttpResponseRedirect
from io import BytesIO

from reportlab.lib.enums import TA_CENTER
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, Table, TableStyle, Image
from reportlab.lib import colors

from sistemaAcademico.Apps.GestionAcademica.Diccionario.Estructuras_tablas_conf import *
from sistemaAcademico.Apps.GestionAcademica.Diccionario.Estructuras_tablas_mant import *

#from trunk.sistemaAcademico.Apps.GestionAcademica.Diccionario.Estructuras_tablas_conf import ConfRol


def view_reporte(request, *args, **kwargs):
    if 'usuario' in request.session:
        if request.method == 'POST':
            usuario = ConfUsuario.objects.get(id_usuario=request.session.get('usuario'))
            print(usuario)
            campoChk = request.POST.get('check1')
            campo2 = request.POST.get('campo')
            combo = int(request.POST.get('combo'))
            comboR = int(request.POST.get('comboR'))
            print('el reporte es: ',comboR)
            if(combo == 1):
                usuarios = ConfUsuario.objects.filter(id_usuario=campo2)
            elif(combo == 2):
                usuarios = ConfUsuario.objects.filter(usuario=campo2)
            elif(combo == 3):
                persona = MantPersona.objects.get(nombres=campo2)
                usuarios = ConfUsuario.objects.filter(id_persona=persona.id_persona)
            if(comboR == 1):
                return reporte_excell(usuarios)
            elif(comboR == 2):
                return reportePdf(usuarios,campoChk,usuario)
        return render(request, 'sistemaAcademico/reportes/reportes.html')
    else:
        return HttpResponseRedirect('timeout/')

def reporte_excell(usuarios):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Hoja' + str()
    # ---------------------------------para darle diseño a mi titulo en la hoja-----------------------------------------

    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws['A1'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))

    ws['A1'].fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type="solid")
    ws['A1'].font = Font(name='Calibri', size=12, bold=True)
    ws['A1'] = 'REPORTE DE USUARIO'
    # ---------------------------------------cambiar caracteristicas de las celdas--------------------------------------
    ws.merge_cells('A1:C1')
    ws.row_dimensions[1].height = 25
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    #ws.column_dimensions['D'].width = 20
    # ----------------------------------------------------darle diseño a mi cabecera------------------------------------
    ws['A3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['A3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['A3'].fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type="solid")
    ws['A3'].font = Font(name='Calibri', size=12, bold=True)
    ws['A3'] = 'Usuario'

    ws['B3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['B3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['B3'].fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type="solid")
    ws['B3'].font = Font(name='Calibri', size=12, bold=True)
    ws['B3'] = 'Nombre'

    ws['C3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['C3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['C3'].fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type="solid")
    ws['C3'].font = Font(name='Calibri', size=12, bold=True)
    ws['C3'] = 'tipo de usuario'
    # ---------------------------pintar datos en excel y EXPORTAR DATOS DE LA BD---------------------------------------------------------
    controlador = 4
    for confusuario in usuarios:
        ws.cell(row=controlador, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=1).border = Border(left=Side(border_style="thin"),
                                                           right=Side(border_style="thin"),
                                                           top=Side(border_style="thin"),
                                                           bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=1).font = Font(name='Calibri', size=8)
        ws.cell(row=controlador, column=1).value = confusuario.usuario

        ws.cell(row=controlador, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=2).border = Border(left=Side(border_style="thin"),
                                                           right=Side(border_style="thin"),
                                                           top=Side(border_style="thin"),
                                                           bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=2).font = Font(name='Calibri', size=8)
        ws.cell(row=controlador,column=2).value = confusuario.id_persona.nombres + " " + confusuario.id_persona.apellidos

        ws.cell(row=controlador, column=3).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=3).border = Border(left=Side(border_style="thin"),
                                                           right=Side(border_style="thin"),
                                                           top=Side(border_style="thin"),
                                                           bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=3).font = Font(name='Calibri', size=8)
        ws.cell(row=controlador, column=3).value = confusuario.id_genr_tipo_usuario.nombre
        controlador += 1
    # cont += 1


    # establecer el nombre de mi archivo
    nombre_archivo = "ReportePersonalizadoExcel.xlsx"
    # Definir tipo de respuesta que va a dar
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename = {0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response



def reportePdf(usuarios,campoChk=None,usuarioph=None):
    response = HttpResponse(content_type='application/pdf')
    #response['Content-Disposition'] = 'attachment; filename=ReportePdf.pdf'
    buffer = BytesIO()
    high = 650
    this_U = None

    styles = getSampleStyleSheet()
    sytlesBH = styles["Heading3"]
    sytlesBH.alignment = TA_CENTER
    sytlesBH.fontSinze = 7

    # styles = getSampleStyleSheet()
    styleN = styles["BodyText"]
    styleN.alignment = TA_CENTER
    styleN.fontSize = 7
    width, height = A4

    usuario = Paragraph('''USUARIO''', sytlesBH)
    persona = Paragraph('''persona''', sytlesBH)
    tipoU = Paragraph('''tipo_usuario''', sytlesBH)
    data = []
    data.append([usuario, persona, tipoU])

    #response['Content-Disposition']='attachment; filename=ReportePdf.pdf'
    for pdf in usuarios:
        this_U= [{'#':pdf.usuario,'h':pdf.id_persona.nombres + " " + pdf.id_persona.apellidos, 'l':pdf.id_genr_tipo_usuario.nombre}]
        #data.append(this_usuario)
        high = high-18


    for students in this_U:
         u = [students['#'], students['h'], students['l']]
         data.append(u)

    c = canvas.Canvas(buffer, pagesize=A4)
    cabecerapdf(high,data,width,height,buffer,c)
    if campoChk != None:
        piePagina(c,usuarioph)

    c.showPage()
    c.save()
    pd = buffer.getvalue()
    buffer.close()
    response.write(pd)
    return response

def cabecerapdf(high,data,width, height,buffer,c):
    c.setLineWidth(.3)
    c.setFont('Helvetica', 22)
    c.drawString(200, 760, 'Reporte Usuario')

    fecha = datetime.datetime.now().date()
    hora = time.strftime("%H:%M")
    c.setFont('Helvetica',10)
    c.drawString(300, 50, 'Fecha en que se genero el reporte: {0}'.format(fecha)+' '+'  hora:{0}'.format(hora))

    c.line(90, 747, 550, 747)
    c.drawImage("static/img/logo-login.png", 50, 760, width=50, height=50)
    table = Table(data, colWidths=[2.9 * cm, 8 * cm, 8.5 * cm])
    table.setStyle(TableStyle([('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),
                               ('BOX', (0, 0), (-1, -1), 0.25, colors.black), ]))
    table.wrapOn(c, width, height)
    table.drawOn(c, 30, high)


def piePagina(c,usuario):
    print(usuario)
    c.setFont('Helvetica',10)
    c.drawString(100, 50, 'Reporte generador por: {0}'.format(usuario))


#//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
def view_Rol(request, *args, **kwargs):
    if 'usuario' in request.session:
        rol = None
        if request.method == 'POST':
            usuario = ConfUsuario.objects.get(id_usuario=request.session.get('usuario'))
            print(usuario)
            campoChk2 = request.POST.get('check2')
            campo2 = request.POST.get('campo')
            if campo2:
                rol = ConfUsuario.objects.filter(id_rol__nombre=campo2)
            else:
                rol = ConfUsuario.objects.all()
            print('kjkjkjkj',campo2)
            comboR = int(request.POST.get('comboR'))
            print('el reporte es: ',comboR)
            if(comboR == 1):
                return reporte_excel_rol(rol)
            elif(comboR == 2):
                return reportePdf_Rol(rol,campoChk2,usuario)
        return render(request, 'sistemaAcademico/reportes/reporterol.html')
    else:
       return HttpResponseRedirect('timeout/')





def reporte_excel_rol(rol):

    wb = Workbook()
    ws = wb.active
    ws.title = 'Hoja' + str()
    # ---------------------------------para darle diseño a mi titulo en la hoja-----------------------------------------

    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws['A1'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))

    ws['A1'].fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type="solid")
    ws['A1'].font = Font(name='Calibri', size=12, bold=True)
    ws['A1'] = 'REPORTE ROL'
    # ---------------------------------------cambiar caracteristicas de las celdas--------------------------------------
    ws.merge_cells('A1:C1')
    ws.row_dimensions[1].height = 25
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    #ws.column_dimensions['D'].width = 20
    # ----------------------------------------------------darle diseño a mi cabecera------------------------------------
    ws['A3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['A3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['A3'].fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type="solid")
    ws['A3'].font = Font(name='Calibri', size=12, bold=True)
    ws['A3'] = 'Nombre del Rol'

    ws['B3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['B3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['B3'].fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type="solid")
    ws['B3'].font = Font(name='Calibri', size=12, bold=True)
    ws['B3'] = 'Usuario'

    ws['C3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['C3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['C3'].fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type="solid")
    ws['C3'].font = Font(name='Calibri', size=12, bold=True)
    ws['C3'] = 'Nombre'

    controlador = 4
    cont = 0
    for confRol in rol:

         ws.cell(row=controlador, column=1).alignment = Alignment(horizontal="center")
         ws.cell(row=controlador, column=1).border = Border(left=Side(border_style="thin"),
                                                            right=Side(border_style="thin"),
                                                            top=Side(border_style="thin"),
                                                            bottom=Side(border_style="thin"))
         ws.cell(row=controlador, column=1).font = Font(name='Calibri', size=8)
         rol_r = [c.nombre for c in confRol.id_rol.all()]
         ws.cell(row=controlador, column=1).value = ', '.join(rol_r)

         ws.cell(row=controlador, column=2).alignment = Alignment(horizontal="center")
         ws.cell(row=controlador, column=2).border = Border(left=Side(border_style="thin"),
                                                            right=Side(border_style="thin"),
                                                            top=Side(border_style="thin"),
                                                            bottom=Side(border_style="thin"))
         ws.cell(row=controlador, column=2).font = Font(name='Calibri', size=8)
         ws.cell(row=controlador,
                 column=2).value = confRol.usuario

         ws.cell(row=controlador, column=3).alignment = Alignment(horizontal="center")
         ws.cell(row=controlador, column=3).border = Border(left=Side(border_style="thin"),
                                                            right=Side(border_style="thin"),
                                                            top=Side(border_style="thin"),
                                                            bottom=Side(border_style="thin"))
         ws.cell(row=controlador, column=3).font = Font(name='Calibri', size=8)
         ws.cell(row=controlador,
                 column=3).value = confRol.id_persona.nombres

         controlador += 1
         cont += 1


    nombre_archivo = "Reporte-Roles.xlsx"
    # Definir tipo de respuesta que va a dar
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename = {0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response


def reportePdf_Rol(rol,campoChk2=None,usuarioph=None):
    response = HttpResponse(content_type='application/pdf')
    #response['Content-Disposition'] = 'attachment; filename=Reporte_Rol.pdf'
    buffer = BytesIO()

    styles = getSampleStyleSheet()
    sytlesBH = styles["Heading3"]
    sytlesBH.alignment = TA_CENTER
    sytlesBH.fontSinze = 7

    # styles = getSampleStyleSheet()
    styleN = styles["BodyText"]
    styleN.alignment = TA_CENTER
    styleN.fontSize = 7
    width, height = A4

    roles = Paragraph('''Nombre del Rol''', sytlesBH)
    usu = Paragraph('''Usuario''', sytlesBH)
    nomb = Paragraph('''Nombre''', sytlesBH)
    data = []
    data.append([roles, usu , nomb])

    #response['Content-Disposition']='attachment; filename=ReportePdf.pdf'
    #contro = 2
    this_rol = []
    for pdfrol in rol:
        miguel = [r.nombre for r in pdfrol.id_rol.all()]
        print(miguel)
        this_rol += [{'R':','.join(miguel) ,'U': pdfrol.usuario, 'N': pdfrol.id_persona.nombres }]
        #this_rol += this_rol

    high = 650
    for Rrol in this_rol:
        u = [Rrol['R'], Rrol['U'], Rrol['N']]
        data.append(u)
        high = high - 18

    c = canvas.Canvas(buffer, pagesize=A4)
    cabecerapdfrol(high,data,width,height,buffer,c)
    if campoChk2 != None:
        piePagina(c,usuarioph)

    c.showPage()
    c.save()
    pd = buffer.getvalue()
    buffer.close()
    response.write(pd)
    return response
def cabecerapdfrol(high,data,width, height,buffer,c):
    c.setLineWidth(.3)
    c.setFont('Helvetica', 22)
    c.drawString(200, 760, 'REPORTE ROL')

    fecha = datetime.datetime.now().date()
    hora = time.strftime("%H:%M")
    c.setFont('Helvetica',10)
    c.drawString(300, 50, 'Fecha en que se genero el reporte: {0}'.format(fecha)+' '+'  hora:{0}'.format(hora))

    c.line(90, 747, 550, 747)
    c.drawImage("static/img/logo-login.png", 50, 760, width=50, height=50)



    table = Table(data, colWidths=[8.5 * cm, 5 * cm, 5 * cm])
    table.setStyle(TableStyle([('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),
                               ('BOX', (0, 0), (-1, -1), 0.25, colors.black), ]))
    table.wrapOn(c, width, height)
    table.drawOn(c, 30, high)



def piePagina(c,usuario):
    print(usuario)
    c.setFont('Helvetica',10)
    c.drawString(100, 50, 'Reporte generador por: {0}'.format(usuario))
#--------------------------------------------------------------------------------------------------------------------------###

def view_mantpersona(request):
    if 'usuario' in request.session:
        persona = None
        if request.method == 'POST':
            usuario = ConfUsuario.objects.get(id_usuario=request.session.get('usuario'))
            print(usuario)
            campoChk = request.POST.get('check1')
            campoP = request.POST.get('campoPersona')
            combo = request.POST.get('combo')
            comboR = int(request.POST.get('comboR'))
            print('el reporte es: ',comboR)
            if(combo == 1):
                persona = MantEstudiante.objects.filter(id_persona__nombres=campoP)
            elif(combo == 2):
                persona = MantEstudiante.objects.filter(tipo_estudiante=campoP)
            elif (combo == 3):
                persona = MantEstudiante.objects.filter(usuario_ing=campoP)
            else:
                persona = MantEstudiante.objects.all()

            if(comboR == 1):
                return mant_estudiante(persona)
            elif(comboR == 2):
                return reportePdf_estudiante(persona,campoChk,usuario)
        return render(request, 'sistemaAcademico/reportes/reportePersona.html')
    else:
        return HttpResponseRedirect('timeout/')



def mant_estudiante(persona):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Hoja' + str()
    # ---------------------------------para darle diseño a mi titulo en la hoja-----------------------------------------

    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws['A1'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))

    ws['A1'].fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type="solid")
    ws['A1'].font = Font(name='Calibri', size=12, bold=True)
    ws['A1'] = 'REPORTE DE ESTUDIANTE'
    # ---------------------------------------cambiar caracteristicas de las celdas--------------------------------------
    ws.merge_cells('A1:C1')
    ws.row_dimensions[1].height = 25
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    #ws.column_dimensions['D'].width = 20
    # ----------------------------------------------------darle diseño a mi cabecera------------------------------------
    ws['A3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['A3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['A3'].fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type="solid")
    ws['A3'].font = Font(name='Calibri', size=12, bold=True)
    ws['A3'] = 'NOMBRE'

    ws['B3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['B3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['B3'].fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type="solid")
    ws['B3'].font = Font(name='Calibri', size=12, bold=True)
    ws['B3'] = 'Tipo Estudiante'

    ws['C3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['C3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['C3'].fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type="solid")
    ws['C3'].font = Font(name='Calibri', size=12, bold=True)
    ws['C3'] = 'Fecha Ingreso'

    ws['D3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['D3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['D3'].fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type="solid")
    ws['D3'].font = Font(name='Calibri', size=12, bold=True)
    ws['D3'] = 'Usuario'
    # ---------------------------pintar datos en excel y EXPORTAR DATOS DE LA BD---------------------------------------------------------
    controlador = 4
    for mant in persona:
        ws.cell(row=controlador, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=1).border = Border(left=Side(border_style="thin"),
                                                           right=Side(border_style="thin"),
                                                           top=Side(border_style="thin"),
                                                           bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=1).font = Font(name='Calibri', size=8)
        ws.cell(row=controlador, column=1).value = mant.id_persona.nombres

        ws.cell(row=controlador, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=2).border = Border(left=Side(border_style="thin"),
                                                           right=Side(border_style="thin"),
                                                           top=Side(border_style="thin"),
                                                           bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=2).font = Font(name='Calibri', size=8)
        ws.cell(row=controlador,column=2).value = mant.tipo_estudiante
        ws.cell(row=controlador, column=3).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=3).border = Border(left=Side(border_style="thin"),
                                                           right=Side(border_style="thin"),
                                                           top=Side(border_style="thin"),
                                                           bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=3).font = Font(name='Calibri', size=8)
        ws.cell(row=controlador, column=3).value = mant.fecha_ingreso

        ws.cell(row=controlador, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=1).border = Border(left=Side(border_style="thin"),
                                                           right=Side(border_style="thin"),
                                                           top=Side(border_style="thin"),
                                                           bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=1).font = Font(name='Calibri', size=8)
        ws.cell(row=controlador, column=1).value = mant.usuario_ing
        controlador += 1
    # cont += 1


    # establecer el nombre de mi archivo
    nombre_archivo = "Reporte.xlsx"
    # Definir tipo de respuesta que va a dar
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename = {0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response




def reportePdf_estudiante(persona,campoChk2=None,usuarioph=None):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=Reporte_estudiante.pdf'
    buffer = BytesIO()

    styles = getSampleStyleSheet()
    sytlesBH = styles["Heading3"]
    sytlesBH.alignment = TA_CENTER
    sytlesBH.fontSinze = 7

    # styles = getSampleStyleSheet()
    styleN = styles["BodyText"]
    styleN.alignment = TA_CENTER
    styleN.fontSize = 7
    width, height = A4

    estu = Paragraph('''Nombre''', sytlesBH)
    testud = Paragraph('''Tipo Estudiante''', sytlesBH)
    datee = Paragraph('''Fecha de ingreso''', sytlesBH)
    usuario = Paragraph('''Usuario''', sytlesBH)
    data = []
    data.append([estu, testud,datee,usuario])

    #response['Content-Disposition']='attachment; filename=ReportePdf.pdf'
    #contro = 2
    this_estudiante = []
    for pdfes in persona:
        this_estudiante += [{'E': pdfes.id_persona.nombres,'S': pdfes.tipo_estudiante, 'T': pdfes.fecha_ingreso, 'U': pdfes.usuario_ing}]
        #this_rol += this_rol

    high = 650
    for estudent in this_estudiante:
        u = [estudent['E'], estudent['S'], estudent['T'], estudent['U']]
        data.append(u)
        high = high - 18

    c = canvas.Canvas(buffer, pagesize=A4)
    cabecerapdf(high,data,width,height,buffer,c)
    if campoChk2 != None:
        piePagina(c,usuarioph)

    c.showPage()
    c.save()
    pd = buffer.getvalue()
    buffer.close()
    response.write(pd)
    return response
def cabecerapdf(high,data,width, height,buffer,c):
    c.setLineWidth(.3)
    c.setFont('Helvetica', 22)
    c.drawString(200, 760, 'REPORTE Estudiante')

    fecha = datetime.datetime.now().date()
    hora = time.strftime("%H:%M")
    c.setFont('Helvetica',10)
    c.drawString(300, 50, 'Fecha: {0}'.format(fecha)+' '+'  hora:{0}'.format(hora))

    c.line(90, 747, 550, 747)
    c.drawImage("static/img/logo-login.png", 50, 760, width=50, height=50)



    table = Table(data, colWidths=[5 * cm, 5 * cm, 5 * cm, 5 * cm])
    table.setStyle(TableStyle([('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),
                               ('BOX', (0, 0), (-1, -1), 0.25, colors.black), ]))
    table.wrapOn(c, width, height)
    table.drawOn(c, 20, high)



def piePagina(c,usuario):
    print(usuario)
    c.setFont('Helvetica',10)
    c.drawString(100, 50, 'Usuario: {0}'.format(usuario))


##############################################################################################################################################

def view_mantempleado(request):
    if 'usuario' in request.session:
        empleado = None
        if request.method == 'POST':
            usuario = ConfUsuario.objects.get(id_usuario=request.session.get('usuario'))
            print(usuario)
            campoChk = request.POST.get('check1')
            campoP = request.POST.get('campoEmpleado')
            combo = request.POST.get('combo')
            comboR = int(request.POST.get('comboR'))
            print('el reporte es: ',comboR)
            if(combo == 1):
                empleado = MantEmpleado.objects.filter(id_persona__nombres=campoP)
            elif(combo == 3):
                empleado = MantEmpleado.objects.filter(id_usuario__usuario=campoP)
            else:
                empleado = MantEmpleado.objects.all()

            if(comboR == 1):
                return mant_empleado(empleado)
            elif(comboR == 2):
                return reportePdf_empleado(empleado,campoChk,usuario)
        return render(request, 'sistemaAcademico/reportes/reporteEmpleado.html')
    else:
        return HttpResponseRedirect('timeout/')



def mant_empleado(empleado):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Hoja' + str()
    # ---------------------------------para darle diseño a mi titulo en la hoja-----------------------------------------

    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws['A1'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))

    ws['A1'].fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type="solid")
    ws['A1'].font = Font(name='Calibri', size=12, bold=True)
    ws['A1'] = 'REPORTE DE Empleado'
    # ---------------------------------------cambiar caracteristicas de las celdas--------------------------------------
    ws.merge_cells('A1:C1')
    ws.row_dimensions[1].height = 25
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    #ws.column_dimensions['D'].width = 20
    # ----------------------------------------------------darle diseño a mi cabecera------------------------------------
    ws['A3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['A3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['A3'].fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type="solid")
    ws['A3'].font = Font(name='Calibri', size=12, bold=True)
    ws['A3'] = 'NOMBRE'

    ws['B3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['B3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['B3'].fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type="solid")
    ws['B3'].font = Font(name='Calibri', size=12, bold=True)
    ws['B3'] = 'Usuario'

    ws['C3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['C3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['C3'].fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type="solid")
    ws['C3'].font = Font(name='Calibri', size=12, bold=True)
    ws['C3'] = 'Fecha Ingreso'

    ws['D3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['D3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['D3'].fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type="solid")
    ws['D3'].font = Font(name='Calibri', size=12, bold=True)
    ws['D3'] = 'Año Electivo'
    # ---------------------------pintar datos en excel y EXPORTAR DATOS DE LA BD---------------------------------------------------------
    controlador = 4
    for empe in empleado:
        ws.cell(row=controlador, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=1).border = Border(left=Side(border_style="thin"),
                                                           right=Side(border_style="thin"),
                                                           top=Side(border_style="thin"),
                                                           bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=1).font = Font(name='Calibri', size=8)
        ws.cell(row=controlador, column=1).value = empe.id_persona_nombres

        ws.cell(row=controlador, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=2).border = Border(left=Side(border_style="thin"),
                                                           right=Side(border_style="thin"),
                                                           top=Side(border_style="thin"),
                                                           bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=2).font = Font(name='Calibri', size=8)
        ws.cell(row=controlador,column=2).value = empe.id_usuario.usuario
        ws.cell(row=controlador, column=3).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=3).border = Border(left=Side(border_style="thin"),
                                                           right=Side(border_style="thin"),
                                                           top=Side(border_style="thin"),
                                                           bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=3).font = Font(name='Calibri', size=8)
        ws.cell(row=controlador, column=3).value = empe.fecha_ingreso

        ws.cell(row=controlador, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=1).border = Border(left=Side(border_style="thin"),
                                                           right=Side(border_style="thin"),
                                                           top=Side(border_style="thin"),
                                                           bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=1).font = Font(name='Calibri', size=8)
        ws.cell(row=controlador, column=1).value = empe.id_anio_lectivo.anio
        controlador += 1
    # cont += 1


    # establecer el nombre de mi archivo
    nombre_archivo = "ReporteEmpleado.xlsx"
    # Definir tipo de respuesta que va a dar
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename = {0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response




def reportePdf_empleado(empleado,campoChk2=None,usuarioph=None):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=Reporte_estudiante.pdf'
    buffer = BytesIO()

    styles = getSampleStyleSheet()
    sytlesBH = styles["Heading3"]
    sytlesBH.alignment = TA_CENTER
    sytlesBH.fontSinze = 7

    # styles = getSampleStyleSheet()
    styleN = styles["BodyText"]
    styleN.alignment = TA_CENTER
    styleN.fontSize = 7
    width, height = A4

    nom = Paragraph('''Nombre''', sytlesBH)
    usua = Paragraph('''Usuario''', sytlesBH)
    datex = Paragraph('''Fecha de ingreso''', sytlesBH)
    anio = Paragraph('''Año Electivo''', sytlesBH)
    data = []
    data.append([nom,usua,datex,anio])

    #response['Content-Disposition']='attachment; filename=ReportePdf.pdf'
    #contro = 2
    this_estudiante = []
    for emp in empleado:
        this_estudiante += [{'E': emp.id_persona.nombres,'M': emp.id_usuario.usuario, 'P': emp.fecha_ingreso, 'L': emp.id_anio_lectivo.anio}]
        #this_rol += this_rol

    high = 650
    for es in this_estudiante:
        u = [es['E'], es['M'], es['P'], es['L']]
        data.append(u)
        high = high - 18

    c = canvas.Canvas(buffer, pagesize=A4)
    cabecerapdfem(high,data,width,height,buffer,c)
    if campoChk2 != None:
        piePagina(c,usuarioph)

    c.showPage()
    c.save()
    pd = buffer.getvalue()
    buffer.close()
    response.write(pd)
    return response
def cabecerapdfem(high,data,width, height,buffer,c):
    c.setLineWidth(.3)
    c.setFont('Helvetica', 22)
    c.drawString(200, 760, 'REPORTE EMPLEADO')

    fecha = datetime.datetime.now().date()
    hora = time.strftime("%H:%M")
    c.setFont('Helvetica',10)
    c.drawString(300, 50, 'Fecha: {0}'.format(fecha)+' '+'  hora:{0}'.format(hora))

    c.line(90, 747, 550, 747)
    c.drawImage("static/img/logo-login.png", 50, 760, width=50, height=50)



    table = Table(data, colWidths=[5 * cm, 5 * cm, 5 * cm, 5 * cm])
    table.setStyle(TableStyle([('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),
                               ('BOX', (0, 0), (-1, -1), 0.25, colors.black), ]))
    table.wrapOn(c, width, height)
    table.drawOn(c, 20, high)



def piePagina(c,usuario):
    print(usuario)
    c.setFont('Helvetica',10)
    c.drawString(100, 50, 'Usuario: {0}'.format(usuario))

