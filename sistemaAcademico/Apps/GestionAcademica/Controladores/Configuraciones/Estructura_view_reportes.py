import openpyxl
from django.shortcuts import render
from django.views.generic import TemplateView
from openpyxl import Workbook #nos permite crear libro de trabajo en excel
import datetime
import time
from openpyxl.styles import Alignment,Border,Font,PatternFill,Side
from django.http.response import HttpResponse, HttpResponseRedirect
from io import BytesIO
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, Table, TableStyle,Image
from reportlab.lib import colors

from sistemaAcademico.Apps.GestionAcademica.Diccionario.Estructuras_tablas_conf import *
from sistemaAcademico.Apps.GestionAcademica.Diccionario.Estructuras_tablas_mant import *
from sistemaAcademico.Apps.GestionAcademica.Diccionario.Estructuras_tablas_mov import *


#from trunk.sistemaAcademico.Apps.GestionAcademica.Diccionario.Estructuras_tablas_conf import ConfRol

from django.template.loader import get_template
from xhtml2pdf import pisa
from datetime import datetime,date
from sistemaAcademico.utils import link_callback


def reporte_usuarios(request, *args, **kwargs):
        if 'usuario' in request.session:
            usuarios = None
            if request.method == 'POST':
                usuario = ConfUsuario.objects.get(id_usuario=request.session.get('usuario'))
                print(usuario)
                campoChk = request.POST.get('check1')
                campo2 = request.POST.get('campo')
                combo = int(request.POST.get('combo'))
                comboR = int(request.POST.get('comboR'))
                print('el reporte es: ', comboR)

                if (combo == 1):
                    usuarios = ConfUsuario.objects.filter(usuario=campo2)
                elif (combo == 2):
                    usuarios = ConfUsuario.objects.filter(id_persona__nombres=campo2)
                elif (combo == 3):
                    usuarios = ConfUsuario.objects.all()
                elif (combo == 5):
                    return render(request, 'sistemaAcademico/reportes/reportes.html')
                    

                if (comboR == 1):
                    return reporte_excell(usuarios, campoChk, usuario)
                elif (comboR == 2):
                    return ReporteUsuario(usuarios, campoChk, usuario)
            return render(request, 'sistemaAcademico/reportes/reportes.html')
        else:
            return HttpResponseRedirect('timeout/')




def reporte_excell(usuarios,campoChk=None,usuarioph=None):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Hoja' + str()

    img = openpyxl.drawing.image.Image('static/img/logo-login.png')
    img.width = 130
    img.height = 65

    # ---------------------------------para darle diseño a mi titulo en la hoja-----------------------------------------
    ws['B2'].alignment = Alignment(horizontal="center", vertical="center")
    ws['B2'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))

    ws['B2'].font = Font(name='times new roman', size=11, bold=True)
    ws['B2'] = 'Fecha:'


    ws['C2'].alignment = Alignment(horizontal="center", vertical="center")
    ws['C2'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['C2'].font = Font(name='times new roman', size=11)
    ws['C2'] = date.today()

    ws['B3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['B3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['B3'].font = Font(name='times new roman', size=11, bold=True)
    ws['B3'] = 'Hora: '

    ws['C3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['C3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['C3'].font = Font(name='times new roman', size=11)
    ws['C3'] = time.strftime("%H:%M")

    ws['B4'].alignment = Alignment(horizontal="center", vertical="center")
    ws['B4'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['B4'].font = Font(name='times new roman', size=11, bold=True)

    ws['C4'].alignment = Alignment(horizontal="center", vertical="center")
    ws['C4'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['C4'].font = Font(name='times new roman', size=11)

    if campoChk != None:
        usur(ws, usuarioph)

    ws['D2'].alignment = Alignment(horizontal="center", vertical="center")
    ws['D2'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"))

    ws['D3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['D3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"))

    ws['D4'].alignment = Alignment(horizontal="center", vertical="center")
    ws['D4'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),bottom=Side(border_style="thin"))


    ws['B5'].alignment = Alignment(horizontal="center", vertical="center")
    ws['B5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))

    ws['B5'].fill = PatternFill(start_color='33FCFF', end_color='33FCFF', fill_type="solid")
    ws['B5'].font = Font(name='times new roman', size=12, bold=True)
    ws['B5'] = 'REPORTE DE USUARIO'
    # ---------------------------------------cambiar caracteristicas de las celdas--------------------------------------
    ws.merge_cells('B5:D5')
    ws.row_dimensions[3].height = 25
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 25

    #ws.column_dimensions['D'].width = 20
    # ----------------------------------------------------darle diseño a mi cabecera------------------------------------

    ws['B6'].alignment = Alignment(horizontal="center", vertical="center")
    ws['B6'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['B6'].fill = PatternFill(start_color='3380FF', end_color='3380FF', fill_type="solid")
    ws['B6'].font = Font(name='times new roman', size=12, bold=True)
    ws['B6'] = 'Usuario'

    ws['C6'].alignment = Alignment(horizontal="center", vertical="center")
    ws['C6'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['C6'].fill = PatternFill(start_color='3380FF', end_color='3380FF', fill_type="solid")
    ws['C6'].font = Font(name='times new roman', size=12, bold=True)
    ws['C6'] = 'Nombre'

    ws['D6'].alignment = Alignment(horizontal="center", vertical="center")
    ws['D6'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['D6'].fill = PatternFill(start_color='3380FF', end_color='3380FF', fill_type="solid")
    ws['D6'].font = Font(name='times new roman', size=12, bold=True)
    ws['D6'] = 'Tipo de Usuario'




    # ---------------------------pintar datos en excel y EXPORTAR DATOS DE LA BD---------------------------------------------------------
    controlador = 7
    for confusuario in usuarios:
        ws.cell(row=controlador, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=2).border = Border(left=Side(border_style="thin"),
                                                           right=Side(border_style="thin"),
                                                           top=Side(border_style="thin"),
                                                           bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=2).font = Font(name='times new roman', size=11)
        ws.cell(row=controlador, column=2).value = confusuario.usuario

        ws.cell(row=controlador, column=3).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=3).border = Border(left=Side(border_style="thin"),
                                                           right=Side(border_style="thin"),
                                                           top=Side(border_style="thin"),
                                                           bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=3).font = Font(name='times new roman', size=11)
        ws.cell(row=controlador,column=3).value = confusuario.id_persona.nombres + " " + confusuario.id_persona.apellidos

        ws.cell(row=controlador, column=4).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=4).border = Border(left=Side(border_style="thin"),
                                                           right=Side(border_style="thin"),
                                                           top=Side(border_style="thin"),
                                                           bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=4).font = Font(name='times new roman', size=11)
        ws.cell(row=controlador, column=4).value = confusuario.id_genr_tipo_usuario.nombre
        controlador += 1


    # establecer el nombre de mi archivo
    nombre_archivo = "ReporteUsuarioExcel.xlsx"
    # Definir tipo de respuesta que va a dar
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename = {0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido

    ws.add_image(img,'D2')
    wb.save('logo-login.xlsx')
    wb.save(response)
    return response
def usur(ws, usuario):
    print(usuario)
    ws['B4'] = 'Usuario: '
    ws['C4'] = ' {0}'.format(usuario)


def ReporteUsuario(usuarios,campoChk=None,usuarioph=None ):
    template_path = 'sistemaAcademico/DiseñoReporte/DiseñoUsuario.html'
    response = HttpResponse(content_type='application/pdf')
    context = {}
    context['fecha_actual'] = date.today()
    context['hora_actual'] = time.strftime("%H:%M")
    response['Content-Disposition'] = 'attachment; filename=ReporteUsuario.pdf'
    context['lista_usuario'] = usuarios
    if campoChk != None: 
        usu(context,usuarioph)

    template = get_template(template_path)
    html = template.render(context)
    pisaStatus = pisa.CreatePDF(html, dest=response, link_callback=link_callback)
    if pisaStatus.err:
        return HttpResponse('we had some errors <pre>'+html+'</pre>')
    return response 
def usu(context,usuario):
    context['nombre_usuario'] = usuario



#//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////



def reporte_roles(request, *args, **kwargs): 
        if 'usuario' in request.session:
            rol = None
            if request.method == 'POST':
                usuario = ConfUsuario.objects.get(id_usuario=request.session.get('usuario'))
                print(usuario)
                campoChk2 = request.POST.get('check2')
                campo2 = request.POST.get('campo')
                combo = int(request.POST.get('combo'))
                if (combo == 1):
                    rol = ConfUsuario.objects.all()
                elif (combo == 2):
                    rol = ConfUsuario.objects.filter(id_rol__nombre=campo2) 
                elif (combo == 3):
                    return render(request, 'sistemaAcademico/reportes/reporterol.html')   
                    

                print('kjkjkjkj', campo2)
                comboR = int(request.POST.get('comboR'))
                print('el reporte es: ', comboR)
                if (comboR == 1):
                    return reporte_excel_rol(rol, campoChk2, usuario)
                elif (comboR == 2):
                    return ReporteRol(rol, campoChk2, usuario)
            return render(request, 'sistemaAcademico/reportes/reporterol.html')
        else:
            return HttpResponseRedirect('timeout/')



def reporte_excel_rol(rol, campoChk2=None,usuarioph=None):

    wb = Workbook()
    ws = wb.active
    ws.title = 'Hoja' + str()

    img = openpyxl.drawing.image.Image('static/img/logo-login.png')
    img.width = 130
    img.height = 65

    # ---------------------------------para darle diseño a mi titulo en la hoja-----------------------------------------
    ws['B2'].alignment = Alignment(horizontal="center", vertical="center")
    ws['B2'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))

    ws['B2'].font = Font(name='times new roman', size=11, bold=True)
    ws['B2'] = 'Fecha:'

    ws['C2'].alignment = Alignment(horizontal="center", vertical="center")
    ws['C2'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['C2'].font = Font(name='times new roman', size=11)
    ws['C2'] = date.today()

    ws['B3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['B3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['B3'].font = Font(name='times new roman', size=11, bold=True)
    ws['B3'] = 'Hora: '

    ws['C3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['C3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['C3'].font = Font(name='times new roman', size=11)
    ws['C3'] = time.strftime("%H:%M")

    ws['B4'].alignment = Alignment(horizontal="center", vertical="center")
    ws['B4'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['B4'].font = Font(name='times new roman', size=11, bold=True)


    ws['C4'].alignment = Alignment(horizontal="center", vertical="center")
    ws['C4'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['C4'].font = Font(name='times new roman', size=11)

    if campoChk2 != None:
        usur(ws,usuarioph)

    ws['D2'].alignment = Alignment(horizontal="center", vertical="center")
    ws['D2'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"))

    ws['D3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['D3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"))

    ws['D4'].alignment = Alignment(horizontal="center", vertical="center")
    ws['D4'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             bottom=Side(border_style="thin"))

    ws['B5'].alignment = Alignment(horizontal="center", vertical="center")
    ws['B5'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))

    ws['B5'].fill = PatternFill(start_color='33FCFF', end_color='33FCFF', fill_type="solid")
    ws['B5'].font = Font(name='times new roman', size=12, bold=True)
    ws['B5'] = 'REPORTE ROL'
    # ---------------------------------------cambiar caracteristicas de las celdas--------------------------------------
    ws.merge_cells('B5:D5')
    ws.row_dimensions[3].height = 25
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 25

    # ws.column_dimensions['D'].width = 20
    # ----------------------------------------------------darle diseño a mi cabecera------------------------------------

    ws['B6'].alignment = Alignment(horizontal="center", vertical="center")
    ws['B6'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['B6'].fill = PatternFill(start_color='3380FF', end_color='3380FF', fill_type="solid")
    ws['B6'].font = Font(name='times new roman', size=12, bold=True)
    ws['B6'] = 'Nombre del Rol'

    ws['C6'].alignment = Alignment(horizontal="center", vertical="center")
    ws['C6'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['C6'].fill = PatternFill(start_color='3380FF', end_color='3380FF', fill_type="solid")
    ws['C6'].font = Font(name='times new roman', size=12, bold=True)
    ws['C6'] = 'Usuario'

    ws['D6'].alignment = Alignment(horizontal="center", vertical="center")
    ws['D6'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['D6'].fill = PatternFill(start_color='3380FF', end_color='3380FF', fill_type="solid")
    ws['D6'].font = Font(name='times new roman', size=12, bold=True)
    ws['D6'] = 'Nombre'

    # ---------------------------pintar datos en excel y EXPORTAR DATOS DE LA BD---------------------------------------------------------
    controlador = 7
    cont = 0
    for confRol in rol:
        ws.cell(row=controlador, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=2).border = Border(left=Side(border_style="thin"),
                                                           right=Side(border_style="thin"),
                                                           top=Side(border_style="thin"),
                                                           bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=2).font = Font(name='times new roman', size=11)
        rol_r = [c.nombre for c in confRol.id_rol.all()]
        ws.cell(row=controlador, column=2).value = ', '.join(rol_r)

        ws.cell(row=controlador, column=3).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=3).border = Border(left=Side(border_style="thin"),
                                                           right=Side(border_style="thin"),
                                                           top=Side(border_style="thin"),
                                                           bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=3).font = Font(name='times new roman', size=11)
        ws.cell(row=controlador,
                column=3).value = confRol.usuario

        ws.cell(row=controlador, column=4).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=4).border = Border(left=Side(border_style="thin"),
                                                           right=Side(border_style="thin"),
                                                           top=Side(border_style="thin"),
                                                           bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=4).font = Font(name='times new roman', size=11)
        ws.cell(row=controlador,
                column=4).value = confRol.id_persona.nombres

        controlador += 1
        cont += 1

    controlador += 1


    # establecer el nombre de mi archivo
    nombre_archivo = "ReporteRolExcel.xlsx"
    # Definir tipo de respuesta que va a dar
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename = {0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido

    ws.add_image(img, 'D2')
    wb.save('logo-login.xlsx')
    wb.save(response)
    return response
def usur(ws, usuario):
    print(usuario)
    ws['B4'] = 'Usuario: '
    ws['C4'] = ' {0}'.format(usuario)




def ReporteRol(rol,campoChk2=None,usuarioph=None):
    template_path = 'sistemaAcademico/DiseñoReporte/DiseñoRol.html'
    response = HttpResponse(content_type='application/pdf')
    context = {}
    context['fecha_actual'] = date.today()
    context['hora_actual'] = time.strftime("%H:%M")
    response['Content-Disposition'] = 'attachment; filename=ReporteRol.pdf'
    context['lista_rol'] = rol
    if campoChk2 != None: 
        rol(context,usuarioph)

    template = get_template(template_path)
    html = template.render(context)
    pisaStatus = pisa.CreatePDF(html, dest=response, link_callback=link_callback)
    if pisaStatus.err:
        return HttpResponse('we had some errors <pre>'+html+'</pre>')
    return response 
def rol(context,usuario):
    context['nombre_usuario'] = usuario


###############################################################################################################
def reporte_horarioEst(request, *args, **kwargs): 
        if 'usuario' in request.session:
            if request.method == 'POST':
                usuario = ConfUsuario.objects.get(id_usuario=request.session.get('usuario'))
                campoChk2 = request.POST.get('check')
                buscador1 = request.POST.get('buscador')
                combo = int(request.POST.get('combo'))
                if (combo == 1):
                    horario = Mov_Horario_materia.objects.all()
                elif (combo == 2):
                    horario = Mov_Horario_materia.objects.filter(id_curso__nombre=buscador1) 

                elif (combo == 3):
                    return render(request, 'sistemaAcademico/reportes/Horario_est.html')   
                    
                comboR = int(request.POST.get('comboR'))
                if (comboR == 2):
                    return HorarioEst(horario, campoChk2, usuario)
                else:
                    return render(request, 'sistemaAcademico/reportes/Horario_est.html')
            return render(request, 'sistemaAcademico/reportes/Horario_est.html')
        else:
            return HttpResponseRedirect('timeout/')



def HorarioEst(horario,campoChk2=None,usuarioph=None ):

    template_path = 'sistemaAcademico/DiseñoReporte/DiseñoHorarioest.html'
    response = HttpResponse(content_type='application/pdf')
    context = {}
    context['fecha_actual'] = date.today()
    context['hora_actual'] = time.strftime("%H:%M")
    response['Content-Disposition'] = 'attachment; filename=HorarioEst.pdf'
    context['horario_estudiante'] = horario
    if campoChk2 != None: 
        horario1(context,usuarioph)

    template = get_template(template_path)
    html = template.render(context)
    pisaStatus = pisa.CreatePDF(html, dest=response, link_callback=link_callback)
    if pisaStatus.err:
        return HttpResponse('we had some errors <pre>'+html+'</pre>')
    return response 
def horario1(context,usuario):
    context['nombre_usuario'] = usuario  


################################################################################################

def reporte_horarioprofe(request, *args, **kwargs): 
        if 'usuario' in request.session:
            if request.method == 'POST':
                usuario = ConfUsuario.objects.get(id_usuario=request.session.get('usuario'))
                campoChk2 = request.POST.get('check')
                buscador1 = request.POST.get('buscador')
                combo = int(request.POST.get('combo'))
                if (combo == 1):
                    horario = Mov_Horario_materia.objects.all()
                elif (combo == 2):
                    horario = Mov_Horario_materia.objects.filter(id_materia_profesor__id_empleado__id_persona__nombres=buscador1) 

                elif (combo == 3):
                    return render(request, 'sistemaAcademico/reportes/Horarioprof.html')   
                    
                comboR = int(request.POST.get('comboR'))
                if (comboR == 2):
                    return HorarioProf(horario, campoChk2, usuario)
                else:
                    return render(request, 'sistemaAcademico/reportes/Horarioprof.html')
            return render(request, 'sistemaAcademico/reportes/Horarioprof.html')
        else:
            return HttpResponseRedirect('timeout/')



def HorarioProf(horario,campoChk2=None,usuarioph=None ):

    template_path = 'sistemaAcademico/DiseñoReporte/DiseñoHorarioprofe.html'
    response = HttpResponse(content_type='application/pdf')
    context = {}
    context['fecha_actual'] = date.today()
    context['hora_actual'] = time.strftime("%H:%M")
    response['Content-Disposition'] = 'attachment; filename=HorarioDocente.pdf'
    context['horario_profesor'] = horario
    if campoChk2 != None: 
        horarios(context,usuarioph)

    template = get_template(template_path)
    html = template.render(context)
    pisaStatus = pisa.CreatePDF(html, dest=response, link_callback=link_callback)
    if pisaStatus.err:
        return HttpResponse('we had some errors <pre>'+html+'</pre>')
    return response 
def horarios(context,usuario):
    context['nombre_usuario'] = usuario
    