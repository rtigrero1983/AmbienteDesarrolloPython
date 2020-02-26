from django.shortcuts import render
from django.views.generic import TemplateView
from openpyxl import Workbook #nos permite crear libro de trabajo en excel
import datetime
import time
from openpyxl.styles import Alignment,Border,Font,PatternFill,Side
from openpyxl.drawing.image import Image
from django.http.response import HttpResponse, HttpResponseRedirect
from io import BytesIO

from reportlab.lib.enums import TA_CENTER
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, Table, TableStyle, Image
from reportlab.lib import colors

from sistemaAcademico.Apps.GestionAcademica.Diccionario.Estructuras_tablas_conf import *

#from trunk.sistemaAcademico.Apps.GestionAcademica.Diccionario.Estructuras_tablas_conf import ConfRol


def view_reporte(request, *args, **kwargs):
    if 'usuario' in request.session:
        if request.method == 'POST':
            usuario = ConfUsuario.objects.get(id_usuario=request.session.get('usuario'))
            print(usuario)
            campoChk = request.POST.get('check1')
            campo2 = request.POST.get('campo')
            combo = int(request.POST.get('combo'))
            comboR = int(request.POST.get('comboR'))
            print('el reporte es: ',comboR)
            if(combo == 1):
                usuarios = ConfUsuario.objects.filter(id_usuario=campo2)
            elif(combo == 2):
                usuarios = ConfUsuario.objects.filter(usuario=campo2)
            elif(combo == 3):
                persona = MantPersona.objects.get(nombres=campo2)
                usuarios = ConfUsuario.objects.filter(id_persona=persona.id_persona)
            if(comboR == 1):
                return reporte_excell(usuarios)
            elif(comboR == 2):
                return reportePdf(usuarios,campoChk,usuario)
        return render(request, 'sistemaAcademico/reportes/reportes.html')
    else:
        return HttpResponseRedirect('timeout/')

def reporte_excell(usuarios):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Hoja' + str()
    # ---------------------------------para darle diseño a mi titulo en la hoja-----------------------------------------

    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws['A1'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))

    ws['A1'].fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type="solid")
    ws['A1'].font = Font(name='Calibri', size=12, bold=True)
    ws['A1'] = 'REPORTE DE USUARIO'
    # ---------------------------------------cambiar caracteristicas de las celdas--------------------------------------
    ws.merge_cells('A1:C1')
    ws.row_dimensions[1].height = 25
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    #ws.column_dimensions['D'].width = 20
    # ----------------------------------------------------darle diseño a mi cabecera------------------------------------
    ws['A3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['A3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['A3'].fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type="solid")
    ws['A3'].font = Font(name='Calibri', size=12, bold=True)
    ws['A3'] = 'Usuario'

    ws['B3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['B3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['B3'].fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type="solid")
    ws['B3'].font = Font(name='Calibri', size=12, bold=True)
    ws['B3'] = 'Nombre'

    ws['C3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['C3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['C3'].fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type="solid")
    ws['C3'].font = Font(name='Calibri', size=12, bold=True)
    ws['C3'] = 'tipo de usuario'
    # ---------------------------pintar datos en excel y EXPORTAR DATOS DE LA BD---------------------------------------------------------
    controlador = 4
    for confusuario in usuarios:
        ws.cell(row=controlador, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=1).border = Border(left=Side(border_style="thin"),
                                                           right=Side(border_style="thin"),
                                                           top=Side(border_style="thin"),
                                                           bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=1).font = Font(name='Calibri', size=8)
        ws.cell(row=controlador, column=1).value = confusuario.usuario

        ws.cell(row=controlador, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=2).border = Border(left=Side(border_style="thin"),
                                                           right=Side(border_style="thin"),
                                                           top=Side(border_style="thin"),
                                                           bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=2).font = Font(name='Calibri', size=8)
        ws.cell(row=controlador,column=2).value = confusuario.id_persona.nombres + " " + confusuario.id_persona.apellidos

        ws.cell(row=controlador, column=3).alignment = Alignment(horizontal="center")
        ws.cell(row=controlador, column=3).border = Border(left=Side(border_style="thin"),
                                                           right=Side(border_style="thin"),
                                                           top=Side(border_style="thin"),
                                                           bottom=Side(border_style="thin"))
        ws.cell(row=controlador, column=3).font = Font(name='Calibri', size=8)
        ws.cell(row=controlador, column=3).value = confusuario.id_genr_tipo_usuario.nombre
        controlador += 1
    # cont += 1


    # establecer el nombre de mi archivo
    nombre_archivo = "ReportePersonalizadoExcel.xlsx"
    # Definir tipo de respuesta que va a dar
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename = {0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response



def reportePdf(usuarios,campoChk=None,usuarioph=None):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=ReportePdf.pdf'
    buffer = BytesIO()
    high = 650




    styles = getSampleStyleSheet()
    sytlesBH = styles["Heading3"]
    sytlesBH.alignment = TA_CENTER
    sytlesBH.fontSinze = 7

    # styles = getSampleStyleSheet()
    styleN = styles["BodyText"]
    styleN.alignment = TA_CENTER
    styleN.fontSize = 7
    width, height = A4

    usuario = Paragraph('''USUARIO''', sytlesBH)
    persona = Paragraph('''persona''', sytlesBH)
    tipoU = Paragraph('''tipo_usuario''', sytlesBH)
    data = []
    data.append([usuario, persona, tipoU])
    #response['Content-Disposition']='attachment; filename=ReportePdf.pdf'
    for pdf in usuarios:
        this_U= [{'#':pdf.usuario,'h':pdf.id_persona.nombres + " " + pdf.id_persona.apellidos, 'l':pdf.id_genr_tipo_usuario.nombre}]
        #data.append(this_usuario)
        #high = high-18


    for students in this_U:
        u = [students['#'], students['h'], students['l']]
        data.append(u)
        high = high-18
    c = canvas.Canvas(buffer, pagesize=A4)
    cabecerapdf(high,data,width,height,buffer,c)
    if campoChk != None:
        piePagina(c,usuarioph)

    c.showPage()
    c.save()
    pd = buffer.getvalue()
    buffer.close()
    response.write(pd)
    return response
def cabecerapdf(high,data,width, height,buffer,c):
    c.setLineWidth(.3)
    c.setFont('Helvetica', 22)
    c.drawString(200, 760, 'Reporte Usuario')

    fecha = datetime.datetime.now().date()
    hora = time.strftime("%H:%M")
    c.setFont('Helvetica',10)
    c.drawString(300, 50, 'Fecha en que se genero el reporte: {0}'.format(fecha)+' '+'  hora:{0}'.format(hora))

    c.line(90, 747, 550, 747)
    c.drawImage("static/img/logo-login.png", 50, 760, width=50, height=50)
    table = Table(data, colWidths=[2.9 * cm, 8 * cm, 8.5 * cm])
    table.setStyle(TableStyle([('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),
                               ('BOX', (0, 0), (-1, -1), 0.25, colors.black), ]))
    table.wrapOn(c, width, height)
    table.drawOn(c, 30, high)


def piePagina(c,usuario):
    print(usuario)
    c.setFont('Helvetica',10)
    c.drawString(100, 50, 'Reporte generador por: {0}'.format(usuario))


#//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
#def view_Rol(request, *args, **kwargs):
#    if 'usuario' in request.session:
#        if request.method == 'POST':
#            usuario = ConfUsuario.objects.get(id_usuario=request.session.get('usuario'))
#            print(usuario)
 #         campoChk = request.POST.get('check1')
 #         campo2 = request.POST.get('campo')
 #         combo = int(request.POST.get('combo'))
 #         comboR = int(request.POST.get('comboR'))
 #         print('el reporte es: ',comboR)
 #         if(combo == 1):
 #             usuarios = ConfUsuario.objects.filter(id_usuario=campo2)
 #         elif(combo == 2):
 #             usuarios = ConfUsuario.objects.filter(usuario=campo2)
 #         elif(combo == 3):
 #             persona = MantPersona.objects.get(nombres=campo2)
 #             usuarios = ConfUsuario.objects.filter(id_persona=persona.id_persona)
 #         if(comboR == 1):
 #             return reporte_excell(usuarios)
 #         elif(comboR == 2):
 #             return reportePdf(usuarios,campoChk,usuario)
 #     return render(request, 'sistemaAcademico/reportes/reportes.html')
 # else:
 #     return HttpResponseRedirect('timeout/')













def reporte_excel_rol(request):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Hoja' + str()
    # ---------------------------------para darle diseño a mi titulo en la hoja-----------------------------------------

    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws['A1'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))

    ws['A1'].fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type="solid")
    ws['A1'].font = Font(name='Calibri', size=12, bold=True)
    ws['A1'] = 'REPORTE Rol'
    # ---------------------------------------cambiar caracteristicas de las celdas--------------------------------------
    ws.merge_cells('A1:C1')
    ws.row_dimensions[1].height = 25
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    #ws.column_dimensions['D'].width = 20
    # ----------------------------------------------------darle diseño a mi cabecera------------------------------------
    ws['A3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['A3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['A3'].fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type="solid")
    ws['A3'].font = Font(name='Calibri', size=12, bold=True)
    ws['A3'] = 'Nombre del Rol'

    ws['B3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['B3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['B3'].fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type="solid")
    ws['B3'].font = Font(name='Calibri', size=12, bold=True)
    ws['B3'] = 'Usuario'

    ws['C3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['C3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                             top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    ws['C3'].fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type="solid")
    ws['C3'].font = Font(name='Calibri', size=12, bold=True)
    ws['C3'] = '#'

    controlador = 4





















    nombre_archivo = "Reporte-Roles.xlsx"
    # Definir tipo de respuesta que va a dar
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename = {0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response